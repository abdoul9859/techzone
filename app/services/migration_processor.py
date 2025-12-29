import asyncio
import json
import csv
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from sqlalchemy.orm import Session
import threading
import time
import requests
import hashlib
import os

from ..database import get_db, Migration, MigrationLog, Product, Client, Supplier
from ..routers.cache import set_cache_item

class MigrationProcessor:
    """Service de traitement des migrations en arrière-plan"""
    
    def __init__(self):
        self.running_migrations: Dict[int, bool] = {}
        self.processing_thread = None
        self.should_stop = False
    
    def start_background_processor(self):
        """Démarre le processeur en arrière-plan"""
        if self.processing_thread is None or not self.processing_thread.is_alive():
            self.should_stop = False
            self.processing_thread = threading.Thread(target=self._background_worker, daemon=True)
            self.processing_thread.start()
            print("✅ Processeur de migrations démarré")
    
    def stop_background_processor(self):
        """Arrête le processeur en arrière-plan"""
        self.should_stop = True
        if self.processing_thread and self.processing_thread.is_alive():
            self.processing_thread.join(timeout=5)
            print("✅ Processeur de migrations arrêté")
    
    def _background_worker(self):
        """Worker en arrière-plan qui traite les migrations"""
        while not self.should_stop:
            try:
                db = next(get_db())
                
                # Chercher les migrations en attente de traitement
                pending_migrations = db.query(Migration).filter(
                    Migration.status == "running",
                    Migration.migration_id.notin_(list(self.running_migrations.keys()))
                ).all()
                
                for migration in pending_migrations:
                    if migration.migration_id not in self.running_migrations:
                        # Marquer comme en cours de traitement
                        self.running_migrations[migration.migration_id] = True
                        
                        # Traiter la migration dans un thread séparé
                        thread = threading.Thread(
                            target=self._process_migration,
                            args=(migration.migration_id,),
                            daemon=True
                        )
                        thread.start()
                
                db.close()
                
            except Exception as e:
                print(f"❌ Erreur dans le worker de migrations: {e}")
            
            time.sleep(2)  # Vérifier toutes les 2 secondes
    
    def _process_migration(self, migration_id: int):
        """Traite une migration spécifique"""
        db = next(get_db())
        
        try:
            migration = db.query(Migration).get(migration_id)
            if not migration:
                return
            
            self._add_log(db, migration_id, "info", f"Début du traitement de la migration: {migration.name}")
            
            # Vérifier si un fichier est associé
            if migration.file_name:
                file_path = Path("uploads") / "migrations" / migration.file_name
                if file_path.exists():
                    self._add_log(db, migration_id, "info", f"Traitement du fichier: {migration.file_name}")
                    
                    # Traiter selon le type de migration
                    success = self._process_file(db, migration, file_path)
                    
                    if success:
                        # Marquer comme terminée avec succès
                        migration.status = "completed"
                        migration.completed_at = datetime.utcnow()
                        self._add_log(db, migration_id, "success", f"Migration terminée avec succès. {migration.success_records} enregistrements traités.")
                    else:
                        # Marquer comme échouée
                        migration.status = "failed"
                        migration.completed_at = datetime.utcnow()
                        migration.error_message = "Erreur lors du traitement du fichier"
                        self._add_log(db, migration_id, "error", "Migration échouée lors du traitement du fichier")
                else:
                    # Fichier non trouvé
                    migration.status = "failed"
                    migration.completed_at = datetime.utcnow()
                    migration.error_message = "Fichier non trouvé"
                    self._add_log(db, migration_id, "error", f"Fichier non trouvé: {migration.file_name}")
            else:
                # Pas de fichier - migration de test
                self._simulate_processing(db, migration)
            
            db.add(migration)
            db.commit()
            
        except Exception as e:
            print(f"❌ Erreur lors du traitement de la migration {migration_id}: {e}")
            migration = db.query(Migration).get(migration_id)
            if migration:
                migration.status = "failed"
                migration.completed_at = datetime.utcnow()
                migration.error_message = str(e)
                self._add_log(db, migration_id, "error", f"Erreur critique: {str(e)}")
                db.add(migration)
                db.commit()
        
        finally:
            # Retirer de la liste des migrations en cours
            if migration_id in self.running_migrations:
                del self.running_migrations[migration_id]
            db.close()
    
    def _process_file(self, db: Session, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier de migration selon son type"""
        try:
            file_extension = file_path.suffix.lower()
            
            if file_extension == '.csv':
                return self._process_csv_file(db, migration, file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return self._process_excel_file(db, migration, file_path)
            elif file_extension == '.json':
                return self._process_json_file(db, migration, file_path)
            else:
                self._add_log(db, migration.migration_id, "error", f"Format de fichier non supporté: {file_extension}")
                return False
                
        except Exception as e:
            self._add_log(db, migration.migration_id, "error", f"Erreur lors du traitement du fichier: {str(e)}")
            return False
    
    def _process_csv_file(self, db: Session, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier CSV"""
        try:
            with open(file_path, 'r', encoding='utf-8', newline='') as csvfile:
                # Détecter le délimiteur
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                rows = list(reader)
                migration.total_records = len(rows)
                
                self._add_log(db, migration.migration_id, "info", f"Fichier CSV chargé: {migration.total_records} lignes")
                
                success_count = 0
                error_count = 0
                
                for index, row in enumerate(rows):
                    try:
                        # Traiter selon le type de migration
                        if migration.type == "products":
                            success = self._import_product_from_row(db, row)
                        elif migration.type == "clients":
                            success = self._import_client_from_row(db, row)
                        elif migration.type == "suppliers":
                            success = self._import_supplier_from_row(db, row)
                        elif migration.type == "invoices":
                            success = self._import_invoice_from_excel_row(db, row)
                        else:
                            success = True  # Migration générique
                        
                        if success:
                            success_count += 1
                        else:
                            error_count += 1
                        
                        # Mettre à jour les compteurs périodiquement
                        if (index + 1) % 10 == 0:
                            migration.processed_records = index + 1
                            migration.success_records = success_count
                            migration.error_records = error_count
                            db.add(migration)
                            db.commit()
                            
                            self._add_log(db, migration.migration_id, "info", f"Progression: {index + 1}/{migration.total_records} lignes traitées")
                    
                    except Exception as e:
                        error_count += 1
                        self._add_log(db, migration.migration_id, "warning", f"Erreur ligne {index + 1}: {str(e)}")
                
                # Mise à jour finale
                migration.processed_records = migration.total_records
                migration.success_records = success_count
                migration.error_records = error_count
                
                return error_count == 0 or success_count > 0
                
        except Exception as e:
            self._add_log(db, migration.migration_id, "error", f"Erreur lors de la lecture du CSV: {str(e)}")
            return False
    
    def _process_excel_file(self, db: Session, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier Excel (nécessite openpyxl)"""
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            # Charger le fichier Excel
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            
            # Lire les en-têtes
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
                else:
                    headers.append(None)
            
            self._add_log(db, migration.migration_id, "info", f"En-têtes détectés: {headers}")
            
            # Compter les lignes de données
            total_rows = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    total_rows += 1
            
            migration.total_records = total_rows
            self._add_log(db, migration.migration_id, "info", f"Fichier Excel chargé: {total_rows} lignes de données")
            
            if total_rows == 0:
                self._add_log(db, migration.migration_id, "warning", "Aucune donnée trouvée dans le fichier Excel")
                return False
            
            success_count = 0
            error_count = 0
            
            # Traiter chaque ligne
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell is not None for cell in row):
                    continue
                
                try:
                    # Créer un dictionnaire avec les données de la ligne
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_data[headers[i]] = value
                    
                    # Traiter selon le type de migration
                    if migration.type == "products":
                        success = self._import_product_from_excel_row(db, row_data)
                    elif migration.type == "clients":
                        success = self._import_client_from_excel_row(db, row_data)
                    elif migration.type == "suppliers":
                        success = self._import_supplier_from_excel_row(db, row_data)
                    elif migration.type == "invoices":
                        success = self._import_invoice_from_excel_row(db, row_data)
                    else:
                        success = True  # Migration générique
                    
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                    
                    # Mettre à jour les compteurs périodiquement
                    if (row_num - 1) % 10 == 0:
                        migration.processed_records = row_num - 1
                        migration.success_records = success_count
                        migration.error_records = error_count
                        db.add(migration)
                        db.commit()
                        
                        self._add_log(db, migration.migration_id, "info", f"Progression: {row_num - 1}/{total_rows} lignes traitées")
                
                except Exception as e:
                    error_count += 1
                    error_msg = f"Erreur ligne {row_num}: {str(e)}"
                    
                    # Rollback de la session en cas d'erreur
                    try:
                        db.rollback()
                    except:
                        pass
                    
                    self._add_log(db, migration.migration_id, "error", error_msg)
            
            # Mise à jour finale
            migration.processed_records = total_rows
            migration.success_records = success_count
            migration.error_records = error_count
            
            workbook.close()
            return error_count == 0 or success_count > 0
            
        except ImportError:
            self._add_log(db, migration.migration_id, "error", "Bibliothèque openpyxl non disponible - installez avec: pip install openpyxl")
            return False
        except Exception as e:
            self._add_log(db, migration.migration_id, "error", f"Erreur lors de la lecture du fichier Excel: {str(e)}")
            return False
    
    def _process_json_file(self, db: Session, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier JSON"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if isinstance(data, list):
                migration.total_records = len(data)
                self._add_log(db, migration.migration_id, "info", f"Fichier JSON chargé: {migration.total_records} enregistrements")
                
                success_count = 0
                error_count = 0
                
                for index, item in enumerate(data):
                    try:
                        # Traiter selon le type
                        if migration.type == "products":
                            success = self._import_product_from_dict(db, item)
                        elif migration.type == "clients":
                            success = self._import_client_from_dict(db, item)
                        elif migration.type == "suppliers":
                            success = self._import_supplier_from_dict(db, item)
                        else:
                            success = True
                        
                        if success:
                            success_count += 1
                        else:
                            error_count += 1
                    
                    except Exception as e:
                        error_count += 1
                        self._add_log(db, migration.migration_id, "warning", f"Erreur enregistrement {index + 1}: {str(e)}")
                
                migration.processed_records = migration.total_records
                migration.success_records = success_count
                migration.error_records = error_count
                
                return error_count == 0 or success_count > 0
            else:
                self._add_log(db, migration.migration_id, "error", "Le fichier JSON doit contenir un tableau")
                return False
                
        except Exception as e:
            self._add_log(db, migration.migration_id, "error", f"Erreur lors de la lecture du JSON: {str(e)}")
            return False
    
    def _simulate_processing(self, db: Session, migration: Migration):
        """Simule le traitement d'une migration sans fichier"""
        migration.total_records = 100
        
        self._add_log(db, migration.migration_id, "info", "Début de la simulation de traitement")
        
        for i in range(0, 101, 10):
            migration.processed_records = i
            migration.success_records = i - (i // 20)  # Quelques erreurs simulées
            migration.error_records = i // 20
            
            db.add(migration)
            db.commit()
            
            self._add_log(db, migration.migration_id, "info", f"Progression: {i}/100 enregistrements traités")
            time.sleep(1)  # Simuler le temps de traitement
        
        migration.status = "completed"
        migration.completed_at = datetime.utcnow()
    
    def _import_product_from_row(self, db: Session, row) -> bool:
        """Importe un produit depuis une ligne CSV/Excel"""
        try:
            # Exemple d'import de produit - à adapter selon votre structure
            product = Product(
                name=str(row.get('name', row.get('nom', ''))),
                description=str(row.get('description', '')),
                price=float(row.get('price', row.get('prix', 0))),
                stock_quantity=int(row.get('stock', row.get('quantite', 0))),
                category_id=1  # Catégorie par défaut
            )
            db.add(product)
            db.commit()
            return True
        except Exception:
            return False
    
    def _import_product_from_excel_row(self, db: Session, row_data: dict) -> bool:
        """Importe un produit depuis une ligne Excel avec structure complète"""
        try:
            from decimal import Decimal
            from datetime import datetime
            
            # Extraire les données avec des noms de colonnes flexibles
            name = self._get_value(row_data, ['name', 'nom', 'product_name', 'produit'])
            if not name:
                return False
            
            description = self._get_value(row_data, ['description', 'desc', 'description_produit'])
            price = self._get_float_value(row_data, ['price', 'prix', 'unit_price', 'prix_unitaire'])
            purchase_price = self._get_float_value(row_data, ['purchase_price', 'prix_achat', 'cost', 'cout'])
            quantity = self._get_int_value(row_data, ['quantity', 'quantite', 'quantité', 'stock', 'qty'])
            category = self._get_value(row_data, ['category', 'categorie', 'catégorie', 'cat'])
            brand = self._get_value(row_data, ['brand', 'marque', 'fabricant'])
            model = self._get_value(row_data, ['model', 'modele', 'modèle', 'reference'])
            barcode = self._get_value(row_data, ['barcode', 'code_barre', 'code-barres', 'ean', 'sku'])
            condition = self._get_value(row_data, ['condition', 'etat', 'state'])
            notes = self._get_value(row_data, ['notes', 'commentaires', 'remarques'])
            image_url = self._get_value(row_data, ['image_path', 'image', 'photo', 'picture', 'img', 'image_url', 'url_image'])
            
            # Si une URL d'image est fournie, la télécharger
            image_path = None
            if image_url:
                image_path = self._download_and_save_image(image_url, name)
            
            # Détecter si c'est un produit avec variantes
            # Chercher des colonnes de variantes (IMEI, série, etc.)
            imei_serial = self._get_value(row_data, ['imei', 'serial', 'imei_serial', 'numéro_série', 'numero_serie'])
            variant_barcode = self._get_value(row_data, ['variant_barcode', 'code_barre_variante', 'barcode_variant'])
            variant_condition = self._get_value(row_data, ['variant_condition', 'condition_variante', 'etat_variante'])
            
            has_variants = bool(imei_serial)  # Si IMEI fourni, c'est une variante
            
            # Valeurs par défaut
            if not price:
                price = 0.0
            if not purchase_price:
                purchase_price = price
            if not quantity:
                quantity = 0
            if not condition:
                condition = "neuf"
            if not variant_condition:
                variant_condition = condition
            
            if has_variants:
                # Produit avec variantes
                return self._import_product_with_variants(
                    db, name, description, price, purchase_price, category, brand, model, 
                    condition, notes, imei_serial, variant_barcode, variant_condition, image_path
                )
            else:
                # Produit sans variantes
                return self._import_simple_product(
                    db, name, description, price, purchase_price, quantity, category, 
                    brand, model, barcode, condition, notes, image_path
                )
            
        except Exception as e:
            return False
    
    def _import_simple_product(self, db: Session, name: str, description: str, price: float, 
                              purchase_price: float, quantity: int, category: str, brand: str, 
                              model: str, barcode: str, condition: str, notes: str, image_path: str = None) -> bool:
        """Importe un produit simple sans variantes"""
        try:
            from decimal import Decimal
            from datetime import datetime
            
            # Créer le produit
            product = Product(
                name=name,
                description=description or "",
                price=Decimal(str(price)),
                purchase_price=Decimal(str(purchase_price)),
                quantity=quantity,
                category=category or "",
                brand=brand or "",
                model=model or "",
                barcode=barcode if barcode else None,
                condition=condition.lower(),
                has_unique_serial=False,
                entry_date=datetime.now(),
                notes=notes or "",
                image_path=image_path if image_path else None
            )
            
            db.add(product)
            db.commit()
            
            # Créer un mouvement de stock d'entrée si quantité > 0
            if quantity > 0:
                from ..database import StockMovement
                stock_movement = StockMovement(
                    product_id=product.product_id,
                    quantity=quantity,
                    movement_type="IN",
                    reference_type="IMPORT_EXCEL",
                    reference_id=None,
                    notes="Import depuis fichier Excel",
                    unit_price=Decimal(str(price))
                )
                db.add(stock_movement)
                db.commit()
            
            return True
            
        except Exception as e:
            return False
    
    def _import_product_with_variants(self, db: Session, name: str, description: str, price: float, 
                                     purchase_price: float, category: str, brand: str, model: str, 
                                     condition: str, notes: str, imei_serial: str, variant_barcode: str, 
                                     variant_condition: str, image_path: str = None) -> bool:
        """Importe un produit avec variantes"""
        try:
            from decimal import Decimal
            from datetime import datetime
            
            # Créer le produit (sans code-barres car il aura des variantes)
            product = Product(
                name=name,
                description=description or "",
                price=Decimal(str(price)),
                purchase_price=Decimal(str(purchase_price)),
                quantity=1,  # 1 variante
                category=category or "",
                brand=brand or "",
                model=model or "",
                barcode=None,  # Pas de code-barres au niveau produit
                condition=condition.lower(),
                has_unique_serial=True,  # Produit avec variantes
                entry_date=datetime.now(),
                notes=notes or "",
                image_path=image_path if image_path else None
            )
            
            db.add(product)
            db.flush()  # Pour obtenir l'ID du produit
            
            # Créer la variante
            from ..database import ProductVariant
            variant = ProductVariant(
                product_id=product.product_id,
                imei_serial=imei_serial,
                barcode=variant_barcode if variant_barcode else None,
                condition=variant_condition.lower(),
                is_sold=False
            )
            
            db.add(variant)
            db.commit()
            
            # Créer un mouvement de stock d'entrée
            from ..database import StockMovement
            stock_movement = StockMovement(
                product_id=product.product_id,
                quantity=1,
                movement_type="IN",
                reference_type="IMPORT_EXCEL",
                reference_id=None,
                notes="Import variante depuis fichier Excel",
                unit_price=Decimal(str(price))
            )
            db.add(stock_movement)
            db.commit()
            
            return True
            
        except Exception as e:
            return False
    
    def _get_value(self, row_data: dict, possible_keys: list) -> str:
        """Récupère une valeur en testant plusieurs clés possibles"""
        for key in possible_keys:
            # Chercher la clé exacte
            value = row_data.get(key)
            if value is not None and str(value).strip():
                return str(value).strip()
            
            # Chercher des variantes avec espaces et accents
            for row_key in row_data.keys():
                if key.lower() in row_key.lower() or row_key.lower() in key.lower():
                    value = row_data.get(row_key)
                    if value is not None and str(value).strip():
                        return str(value).strip()
        return ""
    
    def _get_float_value(self, row_data: dict, possible_keys: list) -> float:
        """Récupère une valeur float en testant plusieurs clés possibles"""
        for key in possible_keys:
            # Chercher la clé exacte
            value = row_data.get(key)
            if value is not None:
                try:
                    return float(value)
                except (ValueError, TypeError):
                    pass
            
            # Chercher des variantes avec espaces et accents
            for row_key in row_data.keys():
                if key.lower() in row_key.lower() or row_key.lower() in key.lower():
                    value = row_data.get(row_key)
                    if value is not None:
                        try:
                            return float(value)
                        except (ValueError, TypeError):
                            pass
        return 0.0
    
    def _download_and_save_image(self, image_url: str, product_name: str) -> Optional[str]:
        """Télécharge une image depuis une URL et la sauvegarde localement"""
        try:
            # Vérifier si c'est une URL valide
            if not image_url.startswith(('http://', 'https://')):
                # Si ce n'est pas une URL, considérer que c'est déjà un chemin local
                return image_url
            
            # Créer le dossier de destination s'il n'existe pas
            upload_dir = Path("static/uploads/products")
            upload_dir.mkdir(parents=True, exist_ok=True)
            
            # Télécharger l'image
            response = requests.get(image_url, timeout=10, stream=True)
            response.raise_for_status()
            
            # Déterminer l'extension du fichier
            content_type = response.headers.get('content-type', '')
            extension = '.jpg'  # Par défaut
            if 'png' in content_type:
                extension = '.png'
            elif 'jpeg' in content_type or 'jpg' in content_type:
                extension = '.jpg'
            elif 'webp' in content_type:
                extension = '.webp'
            elif 'gif' in content_type:
                extension = '.gif'
            
            # Générer un nom de fichier unique basé sur le nom du produit et un hash
            safe_name = "".join(c for c in product_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_name = safe_name.replace(' ', '_')[:50]  # Limiter la longueur
            timestamp = int(datetime.now().timestamp())
            url_hash = hashlib.md5(image_url.encode()).hexdigest()[:8]
            filename = f"{safe_name}_{timestamp}_{url_hash}{extension}"
            
            # Chemin complet du fichier
            file_path = upload_dir / filename
            
            # Sauvegarder l'image
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            # Retourner le chemin relatif pour la base de données
            return f"static/uploads/products/{filename}"
            
        except requests.exceptions.RequestException as e:
            print(f"❌ Erreur lors du téléchargement de l'image {image_url}: {e}")
            return None
        except Exception as e:
            print(f"❌ Erreur lors de la sauvegarde de l'image: {e}")
            return None
    
    def _get_int_value(self, row_data: dict, possible_keys: list) -> int:
        """Récupère une valeur int en testant plusieurs clés possibles"""
        for key in possible_keys:
            # Chercher la clé exacte
            value = row_data.get(key)
            if value is not None:
                try:
                    return int(float(value))  # Convertir via float pour gérer "1.0"
                except (ValueError, TypeError):
                    pass
            
            # Chercher des variantes avec espaces et accents
            for row_key in row_data.keys():
                if key.lower() in row_key.lower() or row_key.lower() in key.lower():
                    value = row_data.get(row_key)
                    if value is not None:
                        try:
                            return int(float(value))  # Convertir via float pour gérer "1.0"
                        except (ValueError, TypeError):
                            pass
        return 0
    
    def _import_client_from_row(self, db: Session, row) -> bool:
        """Importe un client depuis une ligne CSV/Excel"""
        try:
            client = Client(
                name=str(row.get('name', row.get('nom', ''))),
                email=str(row.get('email', '')),
                phone=str(row.get('phone', row.get('telephone', ''))),
                address=str(row.get('address', row.get('adresse', '')))
            )
            db.add(client)
            db.commit()
            return True
        except Exception:
            return False
    
    def _import_supplier_from_row(self, db: Session, row) -> bool:
        """Importe un fournisseur depuis une ligne CSV/Excel"""
        try:
            supplier = Supplier(
                name=str(row.get('name', row.get('nom', ''))),
                email=str(row.get('email', '')),
                phone=str(row.get('phone', row.get('telephone', ''))),
                address=str(row.get('address', row.get('adresse', '')))
            )
            db.add(supplier)
            db.commit()
            return True
        except Exception:
            return False
    
    def _import_product_from_dict(self, db: Session, data: dict) -> bool:
        """Importe un produit depuis un dictionnaire JSON"""
        return self._import_product_from_row(db, data)
    
    def _import_client_from_dict(self, db: Session, data: dict) -> bool:
        """Importe un client depuis un dictionnaire JSON"""
        return self._import_client_from_row(db, data)
    
    def _import_supplier_from_dict(self, db: Session, data: dict) -> bool:
        """Importe un fournisseur depuis un dictionnaire JSON"""
        return self._import_supplier_from_row(db, data)
    
    def _import_invoice_from_excel_row(self, db: Session, row_data: dict) -> bool:
        """Importe une facture ou un article de facture depuis une ligne Excel/CSV"""
        try:
            from decimal import Decimal
            from datetime import datetime
            from ..database import Invoice, InvoiceItem, Client

            # Extraire les champs
            inv_num = self._get_value(row_data, ['invoice_number', 'numero_facture', 'n_facture', 'facture', 'reference', 'ref'])
            if not inv_num:
                return False
            
            client_name = self._get_value(row_data, ['client_name', 'client', 'nom_client', 'nom'])
            if not client_name:
                client_name = "Client Inconnu (Import)"

            # Gestion de la date
            date_val = row_data.get('date')
            inv_date = datetime.now()
            if isinstance(date_val, datetime):
                inv_date = date_val
            elif isinstance(date_val, (int, float)):
                # Gérer les dates Excel (nombre de jours depuis 1900)
                try:
                    from datetime import timedelta
                    inv_date = datetime(1899, 12, 30) + timedelta(days=date_val)
                except:
                    pass
            elif date_val:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                    try:
                        inv_date = datetime.strptime(str(date_val), fmt)
                        break
                    except ValueError:
                        continue

            # Informations produit
            prod_name = self._get_value(row_data, ['product_name', 'produit', 'article', 'description', 'designation'])
            qty = self._get_int_value(row_data, ['quantity', 'quantite', 'qty', 'qte'])
            price = self._get_float_value(row_data, ['price', 'prix', 'unit_price', 'pu'])
            
            # Vérifier si la facture existe déjà
            invoice = db.query(Invoice).filter(Invoice.invoice_number == inv_num).first()
            
            if not invoice:
                # Trouver ou créer le client
                client = db.query(Client).filter(Client.name == client_name).first()
                if not client:
                    client = Client(name=client_name)
                    db.add(client)
                    db.flush()
                
                # Créer la facture
                invoice = Invoice(
                    invoice_number=inv_num,
                    client_id=client.client_id,
                    date=inv_date,
                    status=self._get_value(row_data, ['status', 'statut', 'etat']) or "payée",
                    payment_method=self._get_value(row_data, ['payment_method', 'paiement', 'mode_paiement']) or "espèces",
                    subtotal=Decimal("0"),
                    tax_rate=Decimal("18.00"),
                    tax_amount=Decimal("0"),
                    total=Decimal("0"),
                    paid_amount=Decimal("0"),
                    remaining_amount=Decimal("0"),
                    notes="Importé depuis Excel"
                )
                db.add(invoice)
                db.flush()

            # Ajouter l'article
            if prod_name and qty > 0:
                item_total = Decimal(str(price)) * Decimal(str(qty))
                item = InvoiceItem(
                    invoice_id=invoice.invoice_id,
                    product_name=prod_name,
                    quantity=qty,
                    price=Decimal(str(price)),
                    total=item_total
                )
                db.add(item)
                
                # Mettre à jour les totaux de la facture
                invoice.subtotal += item_total
                # Recalculer la taxe et le total
                invoice.tax_amount = (invoice.subtotal * Decimal(str(invoice.tax_rate)) / 100).quantize(Decimal("1."))
                invoice.total = invoice.subtotal + invoice.tax_amount
                
                # Si le statut est payé, mettre à jour le montant payé
                if invoice.status in ["payée", "payé", "paid"]:
                    invoice.paid_amount = invoice.total
                    invoice.remaining_amount = Decimal("0")
                else:
                    invoice.remaining_amount = invoice.total - invoice.paid_amount
                
                db.add(invoice)
            
            db.commit()
            return True
        except Exception as e:
            db.rollback()
            print(f"❌ Erreur import facture: {e}")
            return False

    def _add_log(self, db: Session, migration_id: int, level: str, message: str):
        """Ajoute un log à une migration"""
        try:
            log = MigrationLog(
                migration_id=migration_id,
                level=level,
                message=message,
                timestamp=datetime.utcnow()
            )
            db.add(log)
            db.commit()
            
            # Mettre en cache pour les performances
            cache_key = f"migration_logs:{migration_id}"
            set_cache_item(cache_key, {"last_log": message, "level": level}, ttl_hours=1, cache_type="migration")
            
        except Exception as e:
            print(f"❌ Erreur lors de l'ajout du log: {e}")

# Instance globale du processeur
migration_processor = MigrationProcessor()
